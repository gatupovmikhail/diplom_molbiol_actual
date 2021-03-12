# Аннотация файла itog.txt путем добавления AD/AR/AD_AR
import openpyxl
wb = openpyxl.load_workbook(filename = 'OMIM-04-2019-AD+AR_mapped.xlsx')
name_sheets = wb.sheetnames
wb.create_sheet(title = 'AD_AR mapped', index = 2)

wb.active = 0
sheet = wb.active
srows = sheet.max_row
scols = sheet.max_column
ad = []
for j in range(2,srows+1):
    cell = sheet.cell(row = j, column = 3)
    ad.append(cell.value)

print(ad)
print(name_sheets)
print(wb.active)
print(srows)

wb.active = 1
sheet = wb.active
srows = sheet.max_row
scols = sheet.max_column
ar = []
for j in range(2,srows+1):
    cell = sheet.cell(row = j, column = 3)
    ar.append(cell.value)

print(ar)
print(name_sheets)
print(wb.active)
print(srows)

# sheet_ad=wb['AD mapped']
# sheet_ar=wb['AR mapped']
# sheet_adar=wb['AD_AR mapped']
# t=1
#
# for k in range(1, sheet_ad.max_column + 1):
#     sheet_adar.cell(row=t, column=k).value = sheet_ad.cell(row=t, column=k).value
#
# for j in range(2,sheet_ad.max_row+1):
#     for i in range(2,sheet_ar.max_row+1):
#         cell_ad = sheet_ad.cell(row=j, column=3)
#         cell_ar = sheet_ar.cell(row=i, column=3)
#         if (cell_ad.value == cell_ar.value):
#             t+=1
#             for k in range(1,sheet_ar.max_column+1):
#                 sheet_adar.cell(row=t,column=k).value=sheet_ar.cell(row=i, column=k).value
#
# wb.save('my_genes.xlsx')
#
#
# wb = openpyxl.load_workbook(filename = 'my_genes.xlsx')
# sheet_adar=wb['AD_AR mapped']
# print('Danger rows:')
# fag=0
# for cell in sheet_adar['F']:
#     if not(('recessive' in cell.value) and ('dominant' in cell.value)):
#         print(cell.row)
#         if not cell.row == 1:
#             fag=1
# if fag==0:
#     print('Keep calm')


### openpyxl.utils.column_index_from_string()   openpyxl.utils.get_column_letter()




###############################################################################################
# name_file='itog.txt'
# #name_file='part_itog.txt'
# out_file=open('itog_annotate.txt','w')
# out_file.close()
# out_file=open('itog_annotate.txt','a')
# with open(name_file) as mfile:
#     zag = mfile.readline()
#     out_file.write(zag[0:-1] + '\t' +'AD/AR' + '\n')
#     for trow in mfile:
#         row=trow.split('\t')
#         flag=0
#         for j in range(len(ad)):
#             if (row[6]==ad[j]):
#                 flag=1
#         for j in range(len(ar)):
#             if (row[6]==ar[j]) and not(flag==1):
#                 flag = 2
#             if (row[6]==ar[j]) and (flag==1):
#                 flag = 3
#
#         if (flag==0):
#             out_file.write(trow[0:-1] + '\t' + '.' + '\n')
#         if (flag==1):
#             out_file.write(trow[0:-1] + '\t' + 'AD' + '\n')
#         if (flag==2):
#             out_file.write(trow[0:-1] + '\t' + 'AR' + '\n')
#         if (flag==3):
#             out_file.write(trow[0:-1] + '\t' + 'AD_AR' + '\n')
#
# out_file.close()
#
#
# print('Данные записаны в ' + 'itog_annotate.txt')